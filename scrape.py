import httpx
import threading
import time
import webbrowser
import os
os.environ["PLAYWRIGHT_BROWSERS_PATH"] = r"C:\Users\kaden\PlaywrightBrowsers"
import datetime
import sys
from urllib.parse import urlencode, quote, quote_plus
import cloudscraper
from bs4 import BeautifulSoup
from parsel import Selector
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
import re
import json
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook
from playwright.sync_api import sync_playwright
from tkinterdnd2 import TkinterDnD, DND_FILES
import mtg_parser

SORTING_MAP = {'best_match': 12}
session = httpx.Client(headers={
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)',
    'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8',
    'Accept-Language': 'en-US,en;q=0.9',
    'Accept-Encoding': 'gzip, deflate, br'
}, http2=True, follow_redirects=True, timeout=30.0)

class RateLimiter:
    def __init__(self, interval: float):
        self.interval = interval
        self.next_time = time.time()

    def wait(self):
        now = time.time()
        if now < self.next_time:
            time.sleep(self.next_time - now)
        self.next_time = time.time() + self.interval

def scrape_moonmtg(query: str):
    import requests, re
    from bs4 import BeautifulSoup
    BASE_URL = "https://moonmtg.com/products/"

    card_name, set_code, number, foil, etched = parse_card_query(query)

    handle = card_name.lower()
    handle = re.sub(r"[’'\":,?!()]", "", handle)
    handle = re.sub(r"[^a-z0-9\s-]", "", handle)
    handle = re.sub(r"\s+", "-", handle)
    handle = handle.strip("-")

    def normalize_variant_title(title: str) -> str:
        t = title.upper()
        t = re.sub(r"\[.*?\]", "", t)
        t = re.sub(r"\(.*?\)", "", t)
        return t.strip()

    try:
        r = requests.get(f"{BASE_URL}{handle}.json", timeout=15)
        if r.status_code != 200:
            return (0.0, "Not found", "")
        product = r.json().get("product", {})
    except Exception:
        return (0.0, "Not found", "")

    variants = product.get("variants", [])
    matches = []

    for v in variants:
        title = v.get("title", "").upper()
        normalized = normalize_variant_title(title)
        vid = v.get("id")
        try:
            price = float(v.get("price", 0))
        except:
            price = 0.0
        if price <= 0:
            continue

        try:
            s = requests.get(f"{BASE_URL}{handle}?variant={vid}", timeout=15)
            if s.status_code != 200:
                continue
            soup = BeautifulSoup(s.text, "html.parser")
            inv = soup.find("p", class_="product__inventory")
            stock_status = inv.get_text(strip=True) if inv else "Unknown"
            if stock_status in ["Out of stock", "Unknown", "Stock info not found"]:
                continue
        except Exception:
            continue

        url = f"{BASE_URL}{handle}?variant={vid}"

        if set_code and number:
            if normalized.startswith(f"{set_code} {number}") or normalized.startswith(f"{set_code}-{number}"):
                if foil and "FOIL" not in title:
                    continue
                if etched and "ETCHED" not in title:
                    continue
                return (price, title, url)
        elif set_code and normalized.startswith(set_code):
            if foil and "FOIL" not in title:
                continue
            if etched and "ETCHED" not in title:
                continue
            matches.append((price, title, url))
        elif not set_code:
            matches.append((price, title, url))

    if matches:
        return min(matches, key=lambda x: x[0])
    return (0.0, "Not found", "")

def parse_card_query(query: str):
    import re
    query = query.strip()
    foil = "*F*" in query
    etched = "*E*" in query
    query = query.replace("*F*", "").replace("*E*", "").strip()

    set_code = None
    number = None

    set_match = re.search(r"\(([a-z0-9]+)\)", query, re.IGNORECASE)
    if set_match:
        set_code = set_match.group(1).upper()
        query = query.replace(set_match.group(0), "").strip()

    id_match = re.search(r"([A-Z0-9]+)-(\d+[a-z]*)", query, re.IGNORECASE)
    if id_match:
        set_code = id_match.group(1).upper()
        number = id_match.group(2)
        query = query.replace(id_match.group(0), "").strip()
    else:
        num_match = re.search(r"(\d+[a-z]*)$", query, re.IGNORECASE)
        if num_match:
            number = num_match.group(1)
            query = query[:num_match.start()].strip()

    card_name = query.strip()
    return card_name, set_code, number, foil, etched

import cloudscraper
from bs4 import BeautifulSoup
import json
import re

def fetch_mtgmate_price(card_name: str, set_name: str = None, set_code: str = None, number: str = None, foil: bool = None):
    url = f"https://www.mtgmate.com.au/cards/search?q={card_name.replace(' ', '+')}"
    scraper = cloudscraper.create_scraper()

    try:
        r = scraper.get(url, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"[DEBUG] Request failed: {e}")
        return (0.0, "Error", "")

    soup = BeautifulSoup(r.text, "html.parser")
    container = soup.find("div", {"data-react-class": "FilterableTable"})
    if not container:
        return (0.0, "Not found", "")

    raw_props = container.get("data-react-props")
    if not raw_props:
        return (0.0, "Not found", "")

    try:
        data = json.loads(raw_props)
    except Exception as e:
        print(f"[DEBUG] JSON parsing error: {e}")
        return (0.0, "Error", "")

    uuid_map = data.get("uuid", {})
    results = []

    target_norm = normalize_name(card_name)

    for card in data.get("cards", []):
        card_id = card.get("uuid")
        details = uuid_map.get(card_id, {})
        if not details:
            continue

        product_name = details.get("name", "")
        product_norm = normalize_name(product_name)

        # ❗ FULL NAME MATCHING — FIX
        if product_norm != target_norm:
            continue

        try:
            price = int(details.get("price", 0)) / 100
        except:
            price = 0.0

        qty = details.get("quantity", 0)
        if price <= 0 or qty <= 0:
            continue

        link_path = details.get("link_path", "")

        match = re.search(r"/([A-Z0-9]+)/(\d+):?", link_path)
        card_set_code = match.group(1) if match else ""
        card_set_number = match.group(2) if match else ""
        card_finish = details.get("finish", "").lower()
        card_set_name = details.get("set_name", "")

        # optional filters (unchanged)
        if set_name and card_set_name.lower() != set_name.lower():
            continue
        if set_code and card_set_code.lower() != set_code.lower():
            continue
        if number and card_set_number != number:
            continue
        if foil is not None and foil != ("foil" in card_finish):
            continue

        results.append((
            price,
            f"{product_name} ({card_set_name}, {details.get('finish')})",
            f"https://www.mtgmate.com.au{link_path}",
        ))

    if not results:
        return (0.0, "Out of stock", "")

    return min(results, key=lambda x: x[0])

def normalize_name(text: str) -> str:
    text = text.lower()
    text = re.sub(r"[’'\":,?!()\[\]{}]", "", text)
    text = re.sub(r"[^a-z0-9\s-]", "", text)
    text = re.sub(r"\s+", " ", text)
    text = text.strip()

    text = text.split("(")[0].split("[")[0].strip()
    return text

def scrape_gg(card_name, base_url):

    def extract_base_name(title):
        title = title.split(" - ")[0]

        title = re.sub(r"\[.*?\]", "", title)
        title = re.sub(r"\(.*?\)", "", title)

        title = title.lower()
        title = re.sub(r"[^a-z0-9\s-]", "", title)
        title = re.sub(r"\s+", " ", title)
        return title.strip()

    target = extract_base_name(card_name)
    query = quote_plus(f"{card_name} product_type:\"mtg\"")
    url = f"{base_url}/search?q={query}"

    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        results = []
        items = soup.select("div.addNow.single")

        for idx, div in enumerate(items, 1):
            onclick = div.get("onclick", "")
            match = re.search(r"addToCart\([^,]+,'([^']+)'", onclick)
            full_title = match.group(1).strip() if match else "N/A"

            price_tag = div.find("p")
            price_text = price_tag.get_text(strip=True) if price_tag else "N/A"
            pm = re.search(r"\$([\d.,]+)", price_text)
            price = float(pm.group(1).replace(",", "")) if pm else 0.0

            title_norm = extract_base_name(full_title)

            if title_norm != target:
                continue

            results.append((price, full_title, url))

        if not results:
            return 0.0, "Out of stock", ""

        return min(results, key=lambda x: x[0])

    except Exception as e:
        return 0.0, "Error", ""

import re
import requests
from bs4 import BeautifulSoup

def scrape_gamesportal(card_name: str):
    import re
    import requests
    from bs4 import BeautifulSoup

    def normalize(text):
        text = text.lower()
        text = re.sub(r"[^a-z0-9\s-]", "", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    target = normalize(card_name)
    url = f"https://gamesportal.com.au/search?type=product&options%5Bprefix%5D=last&q={card_name.replace(' ', '+')}"
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        results = []

        for card in soup.select("div.product-card-list2"):
            title_tag = card.select_one(".grid-view-item__title")
            if not title_tag:
                continue
            title = title_tag.get_text(strip=True)
            if normalize(title.split("(")[0].split("[")[0]) != target:
                continue

            link_tag = card.select_one("a[href]")
            link = link_tag["href"] if link_tag else ""
            if link and not link.startswith("http"):
                link = "https://gamesportal.com.au" + link

            if card.select_one(".outstock-overlay"):
                continue
            if "grid-view-item--sold-out" in " ".join(card.get("class", [])):
                continue

            options = card.select("select.product-form__variants option")
            if options:
                all_disabled = all(
                    opt.has_attr("disabled") or opt.get("data-available") == "0"
                    for opt in options
                )
                if all_disabled:
                    continue

            price_tag = card.select_one(".product-price__price")
            if not price_tag:
                continue
            price_match = re.search(r"\$([\d.,]+)", price_tag.get_text())
            if not price_match:
                continue
            price = float(price_match.group(1).replace(",", ""))

            results.append((price, title, link))

        if not results:
            return 0.0, "Out of stock", ""

        return min(results, key=lambda x: x[0])

    except Exception:
        return 0.0, "Error", ""


def scrape_cardhub(card_name: str):
    import re
    import requests
    from bs4 import BeautifulSoup

    def normalize(text):
        text = text.lower()
        text = re.sub(r"[^a-z0-9\s-]", "", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    target = normalize(card_name)
    url = f"https://thecardhubaustralia.com.au/search?type=product&options%5Bprefix%5D=last&q={card_name.replace(' ', '+')}"
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        results = []

        for card in soup.select("div.product-card-list2"):
            title_tag = card.select_one(".grid-view-item__title")
            if not title_tag:
                continue
            title = title_tag.get_text(strip=True)
            if normalize(title.split("(")[0].split("[")[0]) != target:
                continue

            link_tag = card.select_one("a[href]")
            link = link_tag["href"] if link_tag else ""
            if link and not link.startswith("http"):
                link = "https://thecardhubaustralia.com.au" + link

            if card.select_one(".outstock-overlay"):
                continue
            if "grid-view-item--sold-out" in " ".join(card.get("class", [])):
                continue

            options = card.select("select.product-form__variants option")
            if options:
                all_disabled = all(
                    opt.has_attr("disabled") or opt.get("data-available") == "0"
                    for opt in options
                )
                if all_disabled:
                    continue

            price_tag = card.select_one(".product-price__price")
            if not price_tag:
                continue
            price_match = re.search(r"\$([\d.,]+)", price_tag.get_text())
            if not price_match:
                continue
            price = float(price_match.group(1).replace(",", ""))

            results.append((price, title, link))

        if not results:
            return 0.0, "Out of stock", ""

        return min(results, key=lambda x: x[0])

    except Exception:
        return 0.0, "Error", ""

def scrape_ggadelaide(card_name: str):
    return scrape_gg(card_name, base_url="https://ggadelaide.com.au")

def scrape_ggmodbury(card_name: str):
    return scrape_gg(card_name, base_url="https://ggmodbury.com.au")

def scrape_ggaustralia(card_name: str):
    import re, requests, json, html
    from bs4 import BeautifulSoup
    from urllib.parse import quote_plus, urljoin

    def normalize(text):
        text = html.unescape(text)
        text = text.lower()
        text = re.sub(r"[’'`]", "", text)
        text = re.sub(r"[^a-z0-9\s-]", " ", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def slugify(text):
        text = html.unescape(text)
        text = text.lower()
        text = re.sub(r"[’'`]", "", text)
        text = re.sub(r"[^a-z0-9]+", "-", text)
        text = re.sub(r"-{2,}", "-", text)
        return text.strip("-")

    def find_matching_bracket(text: str, open_pos: int) -> int:
        n = len(text)
        if open_pos < 0 or open_pos >= n or text[open_pos] != "{":
            return -1
        depth = 0
        in_str = False
        str_char = None
        esc = False
        for i in range(open_pos, n):
            ch = text[i]
            if esc:
                esc = False
                continue
            if ch == "\\" and in_str:
                esc = True
                continue
            if ch in ('"', "'"):
                if not in_str:
                    in_str = True
                    str_char = ch
                elif ch == str_char:
                    in_str = False
                    str_char = None
                continue
            if not in_str:
                if ch == "{":
                    depth += 1
                elif ch == "}":
                    depth -= 1
                    if depth == 0:
                        return i
        return -1

    target = normalize(card_name)
    query = quote_plus(card_name)
    search_url = f"https://tcg.goodgames.com.au/search?q={query}&f_Product%20Type=mtg+single"
    headers = {"User-Agent": "Mozilla/5.0"}

    try:
        r = requests.get(search_url, headers=headers, timeout=15)
        r.raise_for_status()
        page_text = r.text
    except Exception as e:
        print(f"[GGAustralia fetch error] {e}")
        return 0.0, "Error", ""

    candidates = []

    key_pattern = re.compile(r"Spurit\.Preorder2\.snippet\.products\[\s*['\"]([^'\"]+)['\"]\s*\]\s*=", re.S)
    for m in key_pattern.finditer(page_text):
        after_eq = m.end()
        brace_pos = page_text.find("{", after_eq)
        if brace_pos == -1:
            continue
        end_pos = find_matching_bracket(page_text, brace_pos)
        if end_pos == -1:
            continue
        block = page_text[brace_pos:end_pos + 1]
        fixed = re.sub(r'([{\s,])([A-Za-z0-9_]+)\s*:', r'\1"\2":', block)
        fixed = fixed.replace("'", '"')
        fixed = re.sub(r',\s*([}\]])', r'\1', fixed)
        try:
            obj = json.loads(fixed)
        except Exception:
            continue

        title = obj.get("title", "")
        base_title = title.split("[")[0].strip()
        normalized_title = normalize(base_title)
        if target != normalized_title:
            continue

        handle = obj.get("handle", "")
        for v in obj.get("variants", []):
            qty = int(v.get("inventory_quantity", 0) or 0)
            if qty <= 0:
                continue
            price_cents = v.get("price")
            if price_cents is None:
                continue
            try:
                price = float(price_cents) / 100.0
            except Exception:
                continue
            if price > 0:
                variant_title = v.get("title", "")
                variant_id = v.get("id")
                product_url = f"https://tcg.goodgames.com.au/products/{handle}"
                if variant_id:
                    product_url += f"?variant={variant_id}"
                candidates.append((price, f"{title} — {variant_title}", product_url))

    if candidates:
        return min(candidates, key=lambda x: x[0])

    try:
        json_url = (
            "https://tcg.goodgames.com.au/search.json?"
            f"q={query}"
            "&f_Product%20Type=mtg+single"
        )
        r2 = requests.get(json_url, headers=headers, timeout=15)
        r2.raise_for_status()
        products = []
        try:
            payload = r2.json()
            products = payload.get("product_data") or payload.get("products") or payload.get("data", {}).get("product_data", [])
        except Exception:
            products = []

        results = []
        for prod in products:
            if str(prod.get("brand", "")).lower() != "magic: the gathering":
                continue
            name = prod.get("name") or ""
            base_name = name.split("[")[0].strip()
            normalized_name = normalize(base_name)
            if target != normalized_name:
                continue
            try:
                price = float(prod.get("price", 0))
            except Exception:
                continue
            if price <= 0:
                continue
            link = f"https://tcg.goodgames.com.au/products/{slugify(name)}"
            results.append((price, name, link))

        if not results:
            return 0.0, "Out of stock", ""
        return min(results, key=lambda x: x[0])

    except Exception as e:
        print(f"[GGAustralia fallback error] {e}")
        return 0.0, "Error", ""

    try:
        soup = BeautifulSoup(page_text, "html.parser")
        results = []
        cards = soup.select("div.search-result, div.grid__item")
        for card in cards:
            title_tag = card.select_one("a.full-unstyled-link, a.product-card__title, a")
            price_tag = card.select_one(".price-item, .price, .product-price")
            if not title_tag or not price_tag:
                continue
            title = title_tag.get_text(strip=True)
            base_title = title.split("[")[0].strip()
            normalized_title = normalize(base_title)
            if target != normalized_title:
                continue
            price_match = re.search(r"\$([\d.,]+)", price_tag.get_text())
            if not price_match:
                continue
            price = float(price_match.group(1).replace(",", ""))
            sold_out = bool(card.find(string=re.compile("Sold out", re.I)))
            if sold_out or price <= 0:
                continue
            link = title_tag.get("href") or "https://tcg.goodgames.com.au"
            if link and not link.startswith("http"):
                link = urljoin("https://tcg.goodgames.com.au", link)
            results.append((price, title, link))
        if not results:
            return 0.0, "Out of stock", ""
        return min(results, key=lambda x: x[0])
    except Exception as e:
        print(f"[GGAustralia DOM fallback error] {e}")
        return 0.0, "Error", ""

def scrape_jenes(card_name: str):
    import requests, re
    from bs4 import BeautifulSoup
    from urllib.parse import quote_plus

    url = f"https://jenesmtg.com.au/search?q={quote_plus(card_name)}&options%5Bprefix%5D=last"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        target = card_name.strip().lower()
        results = []

        for card in soup.select("div.mtg-card"):
            # Skip out of stock
            stock_badge = card.select_one("span.mtg-stock-badge")
            if not stock_badge or "in-stock" not in stock_badge.get("class", []):
                continue

            # Get card name from the anchor title or text
            name_tag = card.select_one("a.mtg-card-name")
            if not name_tag:
                continue

            # Title attr is "CardName|Set|Number | Variant" - take just the card name part
            title_attr = name_tag.get("title", "")
            card_title = title_attr.split("|")[0].strip() if title_attr else name_tag.get_text(strip=True)

            if card_title.lower() != target:
                continue

            link = name_tag.get("href", "")
            if link and not link.startswith("http"):
                link = "https://jenesmtg.com.au" + link
            # Strip tracking params
            link = link.split("?")[0]

            price_tag = card.select_one("span.mtg-card-price")
            if not price_tag:
                continue
            price_text = price_tag.get_text(strip=True)
            match = re.search(r"\$([0-9]+\.[0-9]{2})", price_text)
            if not match:
                continue

            price = float(match.group(1))
            label = title_attr if title_attr else card_title
            results.append((price, label, link))

        if not results:
            return (0.0, "Out of stock", "")

        return min(results, key=lambda x: x[0])

    except Exception as e:
        print(f"[Jene's scrape error]: {e}")
        return (0.0, "Error", "")
    
def scrape_shuffled(card_name: str):
    import requests, re
    from bs4 import BeautifulSoup
    from urllib.parse import quote_plus

    url = f"https://shuffled.com.au/search?page=1&q=%2A{quote_plus(card_name)}%2A"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        target = card_name.strip().lower()
        results = []

        for card in soup.select("div.productCard__card"):
            # Get card title from the link text
            title_tag = card.select_one("p.productCard__title a")
            if not title_tag:
                continue

            full_title = title_tag.get_text(strip=True)
            # Title format: "Abrade (FDN-188) - Foundations"
            # Extract just the card name (before the first bracket or dash)
            card_title = re.split(r"[\(\-]", full_title)[0].strip().lower()
            if card_title != target:
                continue

            # Build product URL
            href = title_tag.get("href", "").split("?")[0]
            link = "https://shuffled.com.au" + href if href.startswith("/") else href

            # Find cheapest available variant from the chip data attributes
            best_price = None
            for chip in card.select("li.productChip"):
                available = chip.get("data-variantavailable", "false") == "true"
                qty = int(chip.get("data-variantqty", "0") or 0)
                if not available or qty <= 0:
                    continue
                try:
                    # Price is stored in cents
                    price_cents = int(chip.get("data-variantprice", "0") or 0)
                    price = price_cents / 100
                except (ValueError, TypeError):
                    continue
                if price > 0 and (best_price is None or price < best_price):
                    best_price = price

            if best_price is not None:
                results.append((best_price, full_title, link))

        if not results:
            return (0.0, "Out of stock", "")

        return min(results, key=lambda x: x[0])

    except Exception as e:
        print(f"[Shuffled scrape error]: {e}")
        return (0.0, "Error", "")

def scrape_kcg(card_name: str):
    import requests, re
    from bs4 import BeautifulSoup
    from urllib.parse import quote_plus

    url = f"https://kastlecardsandgames.com/search?type=product&q={quote_plus(card_name)}"
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}

    try:
        response = requests.get(url, headers=headers, timeout=15)
        response.raise_for_status()
        soup = BeautifulSoup(response.text, "html.parser")

        target = card_name.strip().lower()
        results = []

        for card in soup.select("product-card"):
            # Title is in the visually-hidden span inside the product-card__link
            title_tag = card.select_one("a.product-card__link span.visually-hidden")
            if not title_tag:
                # fallback to h3
                title_tag = card.select_one("h3")
            if not title_tag:
                continue

            full_title = title_tag.get_text(strip=True)
            # Title format: "Sol Ring [Buy-A-Box Promos]" - strip set info in brackets
            card_title = re.split(r"\[", full_title)[0].strip().lower()
            if card_title != target:
                continue

            # Skip sold out cards - badge with "Sold out" text is present
            sold_out_badge = card.select_one("div.product-badges__badge")
            if sold_out_badge and "Sold out" in sold_out_badge.get_text():
                continue

            # Also check for disabled add-to-cart button as a fallback
            add_btn = card.select_one("button.quick-add__button--add")
            if add_btn and add_btn.has_attr("disabled"):
                continue

            # Get product URL from the main link
            link_tag = card.select_one("a.product-card__link")
            href = link_tag.get("href", "").split("?")[0] if link_tag else ""
            link = "https://kastlecardsandgames.com" + href if href.startswith("/") else href

            # Price is in span.price
            price_tag = card.select_one("span.price")
            if not price_tag:
                continue
            price_text = price_tag.get_text(strip=True)
            match = re.search(r"\$([0-9]+\.[0-9]{2})", price_text)
            if not match:
                continue

            price = float(match.group(1))
            results.append((price, full_title, link))

        if not results:
            return (0.0, "Out of stock", "")

        return min(results, key=lambda x: x[0])

    except Exception as e:
        print(f"[KCG scrape error]: {e}")
        return (0.0, "Error", "")

def scrape_hareruyamtg(card_name: str, language_filter: str = "EN") -> tuple:
    """
    Search Hareruya for an MTG single and return the cheapest in-stock match as
    (price_aud, label, url, price_jpy).

    First tries NM stock from the search API. If NM is out of stock, fetches the
    product detail page and scrapes the 'tableHere' div for SP/MP/HP conditions.

    language_filter:
        "EN"    -> English only
        "EN>JP" -> English first; if no EN in stock falls back to cheapest JP
        "JP"    -> Japanese only
        "Other" -> all languages except EN and JP
        "All"   -> no language filter

    Falls back to (0.0, reason, "", 0) on any failure.
    """
    import requests, re
    from bs4 import BeautifulSoup

    HARERUYA_USER_TOKEN = "cc567a4aa1774b15fc1d2a4d94e5bc01fbb701c3f4c8e28085ba8a4661ec3867"
    JPY_TO_AUD = 1 / 113.49

    HEADERS = {
        "Accept": "*/*",
        "Accept-Language": "en-US,en;q=0.8",
        "Referer": "https://www.hareruyamtg.com/en/products/search",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/145.0.0.0 Safari/537.36"
        ),
        "X-Requested-With": "XMLHttpRequest",
        "sec-fetch-dest": "empty",
        "sec-fetch-mode": "cors",
        "sec-fetch-site": "same-origin",
    }

    PAGE_HEADERS = {
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
        "Accept-Language": "en-US,en;q=0.9",
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/145.0.0.0 Safari/537.36"
        ),
    }

    def normalize(text):
        text = text.lower()
        text = re.sub(r"[\u2018\u2019\u0027\":,?!()\[\]{}]", "", text)
        text = re.sub(r"[^a-z0-9\s\-]", "", text)
        text = re.sub(r"\s+", " ", text)
        return text.strip()

    def scrape_conditions_from_page(product_id: str, base_label: str, product_url: str, lang_code: str = "EN") -> list:
        """
        Fetch the product detail page and parse the language-specific priceTable
        div for all in-stock conditions. Returns list of entry tuples.

        In-stock rows have an addCart button; out-of-stock rows have a notifyme button.
        The correct language table is #priceTable-EN or #priceTable-JP.
        """
        try:
            r = requests.get(product_url, headers=PAGE_HEADERS, timeout=20)
            r.raise_for_status()
        except Exception as e:
            print(f"[Hareruya] Detail page fetch error for {product_id}: {e}")
            return []

        soup = BeautifulSoup(r.text, "html.parser")

        # Target the language-specific price table
        table_id = f"priceTable-{lang_code}"
        price_table = soup.select_one(f"#{table_id}")
        if not price_table:
            print(f"[Hareruya] #{table_id} not found on detail page for product {product_id}")
            return []

        entries = []
        for row in price_table.select("div.row.not-first"):
            # Skip the header row (contains "Condition" link text)
            if row.select_one("a[href='/en/user_data/card_condition']"):
                continue

            # Only process rows that have an addCart button (in stock)
            cart_btn = row.select_one("button.addCart.detail")
            if not cart_btn:
                continue

            # Condition: strong tag inside the productClassChange link
            cond_tag = row.select_one("a.productClassChange strong")
            if not cond_tag:
                continue
            cond = cond_tag.get_text(strip=True).upper()

            # Price: col-xs-3 contains e.g. "1,200 JPY"
            price_tag = row.select_one("div.col-xs-3")
            if not price_tag:
                continue
            price_match = re.search(r"[\d,]+", price_tag.get_text(strip=True))
            if not price_match:
                continue
            try:
                price_jpy = float(price_match.group().replace(",", ""))
            except ValueError:
                continue
            if price_jpy <= 0:
                continue

            # Stock: col-xs-2 contains the quantity number
            stock_tag = row.select_one("div.col-xs-2")
            if not stock_tag:
                continue
            try:
                stock = int(stock_tag.get_text(strip=True))
            except ValueError:
                continue
            if stock <= 0:
                continue

            price_aud = round(price_jpy * JPY_TO_AUD, 2)
            label = f"{base_label} [{cond}]"
            entries.append((price_aud, label, product_url, int(price_jpy)))

        return entries

    base_url = "https://www.hareruyamtg.com/en/products/search/unisearch_api"
    params = {
        "kw": card_name,
        "fq.price": "1~*",
        "rows": 60,
        "page": 1,
        "user": HARERUYA_USER_TOKEN,
    }

    # Fetch all pages from search API
    docs = []
    page = 1
    while True:
        params["page"] = page
        try:
            r = requests.get(base_url, params=params, headers=HEADERS, timeout=20)
            r.raise_for_status()
        except requests.RequestException as e:
            print(f"[Hareruya] Request error: {e}")
            return (0.0, "Error", "", 0)
        try:
            data = r.json()
        except ValueError as e:
            print(f"[Hareruya] JSON parse error: {e}")
            return (0.0, "Error", "", 0)

        page_docs = data.get("response", {}).get("docs", [])
        if not page_docs:
            break
        docs.extend(page_docs)
        total = int(data.get("response", {}).get("numFound", 0))
        if len(docs) >= total:
            break
        page += 1

    if not docs:
        return (0.0, "Not found", "", 0)

    target = normalize(card_name)

    # Buckets for NM in-stock entries (from API) and OOS product pages to fall back to
    en_candidates = []
    jp_candidates = []
    other_candidates = []

    # Track out-of-stock NM items so we can scrape their detail pages for SP/MP/HP
    oos_fallback_en = []   # (product_id, base_label, product_url)
    oos_fallback_jp = []
    oos_fallback_other = []

    for item in docs:
        lang_str = str(item.get("language", ""))

        item_name = item.get("card_name") or ""
        if normalize(item_name) != target:
            continue

        # Skip foils and promo/special treatments
        if str(item.get("foil_flg", "0")) == "1":
            continue
        prod_name = (item.get("product_name_en") or item.get("product_name") or "").lower()
        if any(x in prod_name for x in ("foil", "promo", "prerelease", "serial", "galaxy", "retro")):
            continue

        try:
            stock = int(item.get("stock", 0))
        except (TypeError, ValueError):
            stock = 0

        try:
            price_jpy = float(item.get("price", 0))
        except (TypeError, ValueError):
            price_jpy = 0.0

        product_id = item.get("product", "")
        product_url = (
            f"https://www.hareruyamtg.com/en/products/detail/{product_id}?lang=EN"
            if product_id else "https://www.hareruyamtg.com/en/products/search"
        )
        label = (item.get("product_name_en") or item.get("product_name") or item_name).strip()

        print(f"[Hareruya DEBUG] product={product_id!r} lang={lang_str!r} stock={stock} price={price_jpy} label={label!r} keys={sorted(item.keys())}")

        if stock > 0 and price_jpy > 0:
            # NM in stock — add directly
            price_aud = round(price_jpy * JPY_TO_AUD, 2)
            entry = (price_aud, label, product_url, int(price_jpy))
            if lang_str == "2":
                en_candidates.append(entry)
            elif lang_str == "1":
                jp_candidates.append(entry)
            else:
                other_candidates.append(entry)
        else:
            # NM out of stock — record for fallback detail-page scrape
            # lang_str "2"=EN, "1"=JP; map to the priceTable lang code
            if product_id:
                page_lang = "EN" if lang_str == "2" else "JP" if lang_str == "1" else "EN"
                fb = (product_id, label, product_url, page_lang)
                if lang_str == "2":
                    oos_fallback_en.append(fb)
                elif lang_str == "1":
                    oos_fallback_jp.append(fb)
                else:
                    oos_fallback_other.append(fb)

    def resolve(nm_list, oos_list):
        """Return nm_list if non-empty, otherwise scrape detail pages from oos_list."""
        if nm_list:
            return nm_list
        results = []
        for product_id, base_label, product_url, page_lang in oos_list:
            results.extend(scrape_conditions_from_page(product_id, base_label, product_url, page_lang))
        return results

    if language_filter == "EN":
        candidates = resolve(en_candidates, oos_fallback_en)
        if not candidates:
            return (0.0, "Out of stock", "", 0)
        return min(candidates, key=lambda x: x[0])

    elif language_filter == "EN>JP":
        candidates = resolve(en_candidates, oos_fallback_en)
        if candidates:
            return min(candidates, key=lambda x: x[0])
        jp = resolve(jp_candidates, oos_fallback_jp)
        if jp:
            best = min(jp, key=lambda x: x[0])
            return (best[0], f"[JP] {best[1]}", best[2], best[3])
        return (0.0, "Out of stock", "", 0)

    elif language_filter == "JP":
        candidates = resolve(jp_candidates, oos_fallback_jp)
        if not candidates:
            return (0.0, "Out of stock", "", 0)
        return min(candidates, key=lambda x: x[0])

    elif language_filter == "Other":
        candidates = resolve(other_candidates, oos_fallback_other)
        if not candidates:
            return (0.0, "Out of stock", "", 0)
        return min(candidates, key=lambda x: x[0])

    else:  # "All"
        en = resolve(en_candidates, oos_fallback_en)
        jp = resolve(jp_candidates, oos_fallback_jp)
        other = resolve(other_candidates, oos_fallback_other)
        all_candidates = en + jp + other
        if not all_candidates:
            return (0.0, "Out of stock", "", 0)
        return min(all_candidates, key=lambda x: x[0])


def parse_decklist_from_input(text):
    cards = []
    for line in text.strip().splitlines():
        line = line.strip()
        if not line:
            continue
        match = re.match(r'(\d+x?\s*)?(.*)', line, re.IGNORECASE)
        if match:
            card_name = match.group(2).strip()
            if card_name:
                cards.append(card_name) 
    return cards


CACHE_FILE = os.path.join(os.path.dirname(__file__), "deck_cache.json")

def load_deck_cache():
    if os.path.exists(CACHE_FILE):
        try:
            with open(CACHE_FILE, "r", encoding="utf-8") as f:
                return json.load(f)
        except:
            return {}
    return {}

def save_deck_cache(cache):
    with open(CACHE_FILE, "w", encoding="utf-8") as f:
        json.dump(cache, f, indent=2)

SCRAPER_CONFIG = {
    "CardHub":    {"enabled": True, "func": scrape_cardhub},
    "GamesPortal":{"enabled": True, "func": scrape_gamesportal},
    "GGAdelaide": {"enabled": True, "func": scrape_ggadelaide},
    "GGAustralia":{"enabled": True, "func": scrape_ggaustralia},
    "GGModbury":  {"enabled": True, "func": scrape_ggmodbury},
    "Hareruya":   {"enabled": True, "func": scrape_hareruyamtg},
    "JenesMTG":   {"enabled": True, "func": scrape_jenes},
    "KCG":        {"enabled": True, "func": scrape_kcg},
    "MoonMTG":    {"enabled": True, "func": scrape_moonmtg},
    "MTGMate":    {"enabled": True, "func": fetch_mtgmate_price},
    "Shuffled":   {"enabled": True, "func": scrape_shuffled},
}

SOURCE_TO_COLUMN = {
    "MoonMTG": "Moon",
    "MTGMate": "MTGMate",
    "CardHub": "CardHub",
    "GamesPortal": "GamesPortal",
    "JenesMTG": "Jenes",
    "GGAustralia": "GGTCG",
    "GGModbury": "GGModbury",
    "GGAdelaide": "GoodGames",
    "Hareruya": "Hareruya",
}

from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, webbrowser, os, datetime, time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook

from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, webbrowser, os, datetime, time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook

from tkinterdnd2 import TkinterDnD, DND_FILES
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading, webbrowser, os, datetime, time
from concurrent.futures import ThreadPoolExecutor
from openpyxl import Workbook

# ── Card Kingdom price cache ──────────────────────────────────────────────────
_ck_price_cache = {}   # card_name_lower -> cheapest NM retail USD
_ck_cache_ready = False

def _load_ck_prices():
    global _ck_price_cache, _ck_cache_ready
    import requests
    try:
        r = requests.get(
            "https://api.cardkingdom.com/api/v2/pricelist",
            headers={"User-Agent": "MTGPriceChecker/1.0 (contact: kadenschaedel@gmail.com)"},
            timeout=30
        )
        r.raise_for_status()
        data = r.json()
        cache = {}
        for item in data.get("data", []):
            if str(item.get("is_foil", "0")) == "1":
                continue
            name = item.get("name", "").strip().lower()
            try:
                price = float(item.get("price_retail", 0) or 0)
            except (TypeError, ValueError):
                continue
            if price <= 0:
                continue
            # Keep cheapest non-foil NM price per card name
            if name not in cache or price < cache[name]:
                cache[name] = price
        _ck_price_cache = cache
        _ck_cache_ready = True
        print(f"[CK] Price list loaded: {len(cache)} cards")
    except Exception as e:
        print(f"[CK] Failed to load price list: {e}")
        _ck_cache_ready = True  # mark ready so UI doesn't wait forever

def get_ck_price(card_name: str):
    """Return cheapest CK NM non-foil USD price, or None if not found."""
    return _ck_price_cache.get(card_name.strip().lower())

class MTGScraperGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("MTG Price Checker")
        self.root.geometry("1500x750")
        self.root.resizable(False, False)
        self.card_urls = {}
        self.stop_flag = False
        self._results_cache = {}  # card_name -> (display_data, cheapest, url, results)
        self.last_selected_row = None
        headers = {"user-agent": "my-mtg-scraper/1.0 (contact: kadenschaedel@gmail.com)"}
        self.http_client = httpx.Client(headers=headers)

        # ── Toolbar ───────────────────────────────────────────────────
        toolbar = tk.Frame(root, bd=1, relief="raised")
        toolbar.pack(side="top", fill="x")

        # Force treeview text to black by default
        style = ttk.Style()
        style.configure("Treeview", foreground="black", background="white", fieldbackground="white")
        style.map("Treeview", foreground=[("selected", "white")], background=[("selected", "#0078d7")])

        self.quick_menu_button = tk.Menubutton(toolbar, text="Open Sources", relief="raised")
        self.quick_menu = tk.Menu(self.quick_menu_button, tearoff=0)
        for source in SCRAPER_CONFIG:
            self.quick_menu.add_command(
                label=f"From {source}",
                command=lambda s=source: self.open_cheapest_from_source(s)
            )
        self.quick_menu.add_command(label="All from All Sources", command=self.open_all_cheapest_by_source)
        self.quick_menu_button.config(menu=self.quick_menu)
        self.quick_menu_button.pack(side="left", padx=5, pady=2)

        self.toggles_button = tk.Menubutton(toolbar, text="Toggles", relief="raised")
        self.toggles_menu = tk.Menu(self.toggles_button, tearoff=0)
        self.source_vars = {}
        for source in SCRAPER_CONFIG:
            var = tk.BooleanVar(value=SCRAPER_CONFIG[source]['enabled'])
            self.source_vars[source] = var
            self.toggles_menu.add_checkbutton(label=source, variable=var, command=self.on_source_toggle)
        self.toggles_button.config(menu=self.toggles_menu)
        self.toggles_button.pack(side="left", padx=5, pady=2)

        # ── Main area: top panel (left+right) + results table ─────────
        # Top panel holds left (deck input) and right (settings/missing)
        top_panel = tk.Frame(root)
        top_panel.pack(side="top", fill="x", padx=6, pady=(4, 0))

        # ── LEFT: Deck Input ──────────────────────────────────────────
        deck_frame = tk.LabelFrame(top_panel, text="Deck Input")
        deck_frame.pack(side="left", fill="both", expand=True, padx=(0, 4), pady=2)

        url_frame = tk.Frame(deck_frame)
        url_frame.pack(fill='x', padx=4, pady=(4, 2))

        self.url_entry = tk.Entry(url_frame, fg="grey", width=32)
        self.url_entry.insert(0, "Paste a deck link")
        self.url_entry.pack(side="left", fill="x", expand=True, padx=(0, 4))
        self.url_entry.bind("<FocusIn>", self.clear_placeholder)
        self.url_entry.bind("<FocusOut>", self.add_placeholder)

        self.fetch_button = tk.Button(url_frame, text="Fetch", command=self.fetch_deck_from_url)
        self.fetch_button.pack(side="left", padx=2)

        self.save_deck_button = tk.Button(url_frame, text="Save Deck", command=self.save_deck)
        self.save_deck_button.pack(side="left", padx=2)

        saved_frame = tk.Frame(deck_frame)
        saved_frame.pack(fill='x', padx=4, pady=2)

        self.deck_var = tk.StringVar()
        self.deck_dropdown = ttk.Combobox(saved_frame, textvariable=self.deck_var, state="readonly", width=30)
        self.deck_dropdown.pack(side="left", padx=(0, 4), fill="x", expand=True)
        self.deck_dropdown.bind("<<ComboboxSelected>>", self.load_saved_deck)
        self.deck_dropdown.set("Select saved deck")
        self.deck_dropdown.configure(foreground="grey")
        self.deck_dropdown.bind("<FocusIn>", self.clear_dropdown_placeholder)
        self.deck_dropdown.bind("<FocusOut>", self.add_dropdown_placeholder)

        self.delete_deck_button = tk.Button(saved_frame, text="Delete", command=self.delete_deck)
        self.delete_deck_button.pack(side="left", padx=2)

        self.deck_cache = load_deck_cache()
        self.refresh_deck_dropdown()

        self.text_input = tk.Text(deck_frame, height=12, width=50, wrap='word', relief="sunken", borderwidth=2)
        self.text_input.pack(pady=(2, 4), padx=4, fill='both', expand=True)
        self.text_input.drop_target_register(DND_FILES)
        self.text_input.dnd_bind("<<Drop>>", self.handle_file_drop)

        btn_row = tk.Frame(deck_frame)
        btn_row.pack(fill="x", padx=4, pady=(0, 4))

        self.button = tk.Button(btn_row, text='Search Prices', width=14, command=self.toggle_search, bg="#4CAF50", fg="white", font=("Helvetica", 10, "bold"))
        self.button.pack(side="left", padx=(0, 4))

        self.load_button = tk.Button(btn_row, text='Load File', command=self.load_file)
        self.load_button.pack(side="left", padx=2)

        self.save_button = tk.Button(btn_row, text='Save to Excel', command=self.save_to_excel)
        self.save_button.pack(side="left", padx=2)

        # ── RIGHT: Settings + Missing Cards ───────────────────────────
        right_panel = tk.Frame(top_panel, width=210)
        right_panel.pack(side="left", fill="y", padx=(0, 0), pady=2)
        right_panel.pack_propagate(False)

        settings_frame = tk.LabelFrame(right_panel, text="Settings")
        settings_frame.pack(fill="x", padx=2, pady=(0, 4))

        # Hareruya language
        lang_row = tk.Frame(settings_frame)
        lang_row.pack(fill="x", padx=6, pady=(6, 2))
        tk.Label(lang_row, text="Hareruya Language:", anchor="w").pack(side="left")
        self.hareruya_lang_var = tk.StringVar(value="EN")
        self.hareruya_lang_dropdown = ttk.Combobox(
            lang_row,
            textvariable=self.hareruya_lang_var,
            values=["EN", "EN>JP", "JP", "Other", "All"],
            state="readonly",
            width=7
        )
        self.hareruya_lang_dropdown.pack(side="left", padx=(4, 0))

        # Sideboard / Maybeboard toggles
        self.include_sideboard = tk.BooleanVar(value=False)
        self.include_maybeboard = tk.BooleanVar(value=False)
        tk.Checkbutton(settings_frame, text="Include Sideboard", variable=self.include_sideboard).pack(anchor="w", padx=6, pady=1)
        tk.Checkbutton(settings_frame, text="Include Maybeboard", variable=self.include_maybeboard).pack(anchor="w", padx=6, pady=(1, 6))

        missing_frame = tk.LabelFrame(right_panel, text="Missing Cards")
        missing_frame.pack(fill="both", expand=True, padx=2, pady=0)

        self.missing_listbox = tk.Listbox(missing_frame, height=8, width=24)
        self.missing_listbox.pack(side='left', fill='both', expand=True, padx=(4, 0), pady=4)
        missing_scroll = ttk.Scrollbar(missing_frame, orient='vertical', command=self.missing_listbox.yview)
        missing_scroll.pack(side='right', fill='y', pady=4, padx=(0, 4))
        self.missing_listbox.config(yscrollcommand=missing_scroll.set)

        # ── Results table ─────────────────────────────────────────────
        table_frame = tk.Frame(root)
        table_frame.pack(side="top", fill='both', expand=True, padx=6, pady=(4, 0))

        self.tree = ttk.Treeview(table_frame, columns=('Card',) + tuple(SCRAPER_CONFIG.keys()) + ('Cheapest', 'CK%'), show='headings')
        for col in self.tree['columns']:
            self.tree.heading(col, text=col, command=lambda c=col: self.sort_treeview(c, False))
            self.tree.column(col, width=100)
        self.tree.column('Card', width=160)
        self.tree.column('CK%', width=70)

        # Configure CK% colour tags
        self._ck_tag_cache = {}
        self.context_menu = tk.Menu(self.tree, tearoff=0)
        for source in SCRAPER_CONFIG.keys():
            self.context_menu.add_command(
                label=f"Open from {source}",
                command=lambda s=source: self.open_from_source(s)
            )

        h_scroll = ttk.Scrollbar(table_frame, orient='horizontal', command=self.tree.xview)
        v_scroll = ttk.Scrollbar(table_frame, orient='vertical', command=self.tree.yview)
        self.tree.configure(yscrollcommand=v_scroll.set, xscrollcommand=h_scroll.set)
        self.tree.grid(row=0, column=0, sticky='nsew')
        v_scroll.grid(row=0, column=1, sticky='ns')
        h_scroll.grid(row=1, column=0, sticky='ew')
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        self.tree.bind('<ButtonRelease-1>', self.on_click)
        self.tree.bind("<Button-3>", self.show_context_menu)

        # Apply initial column visibility
        self.update_visible_columns()

        # Load CK prices in background
        threading.Thread(target=_load_ck_prices, daemon=True).start()

        # ── Status bar ────────────────────────────────────────────────
        status_bar = tk.Frame(root, bd=1, relief="sunken")
        status_bar.pack(side="bottom", fill="x", padx=6, pady=4)

        self.progress_label = tk.Label(status_bar, text='Ready', anchor="w")
        self.progress_label.pack(side="left", padx=8)

        self.total_label = tk.Label(status_bar, text='Total: AU $0.00', font=('Helvetica', 11, 'bold'), anchor="center")
        self.total_label.pack(side="left", expand=True)

        self.open_all_button = tk.Button(status_bar, text='Open All Cheapest', command=self.open_all_cheapest)
        self.open_all_button.pack(side="right", padx=8, pady=2)

    def show_context_menu(self, event):
        selected = self.tree.identify_row(event.y)
        if selected:
            self.tree.selection_set(selected) 
            self.context_menu.tk_popup(event.x_root, event.y_root)


    def sort_treeview(self, col, reverse=False):
        rows = [(self.tree.set(k, col), k) for k in self.tree.get_children('')]
    
        def try_float(val):
            try:
                return float(val)
            except:
                return val.lower() if isinstance(val, str) else val

        rows.sort(key=lambda t: try_float(t[0]), reverse=reverse)

        for index, (val, k) in enumerate(rows):
            self.tree.move(k, '', index)

        self.tree.heading(col, command=lambda: self.sort_treeview(col, not reverse))

    def clear_placeholder(self, event=None):
        if self.url_entry.get() == "Paste a deck link":
            self.url_entry.delete(0, tk.END)
            self.url_entry.config(fg="black")

    def add_placeholder(self, event=None):
        if not self.url_entry.get().strip():
            self.url_entry.delete(0, tk.END) 
            self.url_entry.insert(0, "Paste a deck link")
            self.url_entry.config(fg="grey")

    def clear_dropdown_placeholder(self, event=None):
        if self.deck_dropdown.get() == "Select saved deck":
            self.deck_dropdown.set("")
            self.deck_dropdown.configure(foreground="black")

    def add_dropdown_placeholder(self, event=None):
        if not self.deck_dropdown.get().strip():
            self.deck_dropdown.set("Select saved deck")
            self.deck_dropdown.configure(foreground="grey")

    def fetch_moxfield_deck(url: str):
        match = re.search(r"/decks/([a-zA-Z0-9\-_]+)", url)
        if not match:
            raise ValueError("Invalid Moxfield URL")
        deck_id = match.group(1)

        api_url = f"https://api.moxfield.com/v2/decks/all/{deck_id}"
        print(f"[DEBUG] Fetching from: {api_url}")

        headers = {
            "User-Agent": "my-mtg-scraper/1.0 (contact: kadenschaedel@gmail.com)"
        }
        r = requests.get(api_url, headers=headers, timeout=15)
        r.raise_for_status()
        data = r.json()

        cards = []
        for section in ("mainboard", "sideboard", "maybeboard"):
            if section in data:
                for card in data[section].values():
                    qty = card["quantity"]
                    name = card["card"]["name"]
                    cards.append((qty, name))
                    print(f"[DEBUG] Parsed {qty}x {name}")
        return cards

    def fetch_deck_from_url(self):
        url = self.url_entry.get().strip()
        if not url or url == "Paste deck link":
            messagebox.showwarning("Input Error", "Please paste a deck link.")
            return

        try:
            import httpx

            if "moxfield.com" in url:
                messagebox.showwarning("Error", "Moxfield not currently supported.")
                return
            else:
                cards = list(mtg_parser.parse_deck(url))

            if not cards:
                messagebox.showerror("Error", "Could not parse decklist from the provided URL.")
                return

            filtered = []
            for c in cards:
                if "sideboard" in c.tags and not self.include_sideboard.get():
                    continue
                if "maybeboard" in c.tags and not self.include_maybeboard.get():
                    continue
                filtered.append(c)

            deck_text = "\n".join([f"{c.quantity} {c.name}" for c in filtered])

            self.text_input.delete('1.0', tk.END)
            self.text_input.insert(tk.END, deck_text)

        except Exception as e:
            import traceback
            traceback.print_exc()
            messagebox.showerror("Error", f"Failed to fetch deck:\n{e}")



    def handle_file_drop(self, event):
        path = event.data.strip("{}") 
        if os.path.isfile(path):
            with open(path, 'r', encoding='utf-8') as f:
                self.text_input.delete('1.0', tk.END)
                self.text_input.insert(tk.END, f.read())

    def update_visible_columns(self):
        """Show only enabled source columns plus Card and Cheapest."""
        visible = ['Card']
        for source in SCRAPER_CONFIG:
            if self.source_vars[source].get():
                visible.append(source)
        visible.append('Cheapest')
        visible.append('CK%')
        self.tree['displaycolumns'] = visible

    def on_source_toggle(self):
        self.update_visible_columns()
        self.recalculate_cheapest_prices()

    def recalculate_cheapest_prices(self):
        total = 0.0
        missing_cards = []
        for row_id in self.tree.get_children():
            values = self.tree.item(row_id)['values']
            card_name = values[0]
            new_row = [card_name]
            cheapest_price = float('inf')
            cheapest_url = ""
            for source in SCRAPER_CONFIG:
                price_str = self.card_urls.get(card_name, {}).get('Prices', {}).get(source, "--")
                try:
                    # Handle "1.32 (¥150)" format from Hareruya
                    price = float(str(price_str).split()[0])
                except:
                    price = 0.0
                if self.source_vars[source].get():
                    new_row.append(price_str if source == "Hareruya" and price > 0 else f"{price:.2f}" if price > 0 else "--")
                    if 0 < price < cheapest_price:
                        cheapest_price = price
                        cheapest_url = self.card_urls.get(card_name, {}).get('URLs', {}).get(source, "")
                else:
                    new_row.append("--")
            if cheapest_price == float('inf'):
                cheapest_price = 0.0
            new_row.append(f"{cheapest_price:.2f}")
            # Recalculate CK% with updated cheapest
            ck_usd = get_ck_price(str(card_name))
            if ck_usd and cheapest_price > 0:
                ratio = ck_usd / cheapest_price
                new_row.append(f"{ratio*100:.0f}%")
            else:
                new_row.append("--")
            self.tree.item(row_id, values=tuple(new_row))
            if ck_usd and cheapest_price > 0:
                self._apply_ck_tag(row_id, ratio)
            self.card_urls[card_name]['Cheapest'] = cheapest_url
            total += cheapest_price
            if cheapest_price == 0.0:
                missing_cards.append(card_name)
        self.total_label.config(text=f"Total: AU ${total:.2f}")
        self.missing_listbox.delete(0, tk.END)
        for card in sorted(set(missing_cards)):
            self.missing_listbox.insert(tk.END, card)

    def _apply_ck_tag(self, iid, ratio):
        """Colour the entire row background based on CK ratio.
        Green = cheaper than CK, red = more expensive, white = same price."""
        clamped = max(0.25, min(2.5, ratio))
        if ratio >= 1.0:
            intensity = min(1.0, (clamped - 1.0) / 1.5)
            r = int(255 - intensity * 120)
            g = 255
            b = int(255 - intensity * 120)
        else:
            intensity = min(1.0, (1.0 - clamped) / 0.75)
            r = 255
            g = int(255 - intensity * 120)
            b = int(255 - intensity * 120)
        colour = f"#{r:02x}{g:02x}{b:02x}"
        tag = f"ck_{colour}"
        if tag not in self._ck_tag_cache:
            self.tree.tag_configure(tag, background=colour, foreground="black")
            self._ck_tag_cache[tag] = True
        existing = [t for t in list(self.tree.item(iid, 'tags') or []) if not t.startswith('ck_')]
        existing.append(tag)
        self.tree.item(iid, tags=existing)

    def toggle_search(self):
        if self.button['text'] == 'Search Prices':
            self.button.config(text='Stop')
            self.hareruya_lang_dropdown.config(state='disabled')
            self.stop_flag = False
            threading.Thread(target=self.check_prices, daemon=True).start()
        else:
            self.stop_flag = True
            self.button.config(state='disabled')

    def load_file(self):
        filepath = filedialog.askopenfilename(filetypes=[('Text files', '*.txt')])
        if filepath:
            with open(filepath, 'r', encoding='utf-8') as f:
                self.text_input.delete('1.0', tk.END)
                self.text_input.insert(tk.END, f.read())

    def fetch_card_prices_parallel(self, card):
        enabled_sources = {name: cfg['func'] for name, cfg in SCRAPER_CONFIG.items() if self.source_vars[name].get()}
        hareruya_lang = self.hareruya_lang_var.get()
        futures = {}
        with ThreadPoolExecutor(max_workers=len(enabled_sources)) as executor:
            for name, func in enabled_sources.items():
                if name == "eBay":
                   continue
                if name == "Hareruya":
                    futures[name] = executor.submit(func, card, hareruya_lang)
                else:
                    futures[name] = executor.submit(func, card)

        results = {}
        for name, future in futures.items():
            try:
                result = future.result()
                if isinstance(result, tuple) and len(result) >= 3:
                    # Normalise to 3-tuple for rest of app; store full result separately
                    results[name] = result[:3]
                    if name == "Hareruya" and len(result) == 4:
                        results[f"{name}_jpy"] = result[3]
                else:
                    results[name] = (0.0, "Invalid result", "")
            except Exception as e:
                print(f"[{name} scrape error]: {e}")
                results[name] = (0.0, "Error", "")

        if "eBay" in enabled_sources:
            time.sleep(0.75)
            results["eBay"] = SCRAPER_CONFIG["eBay"]["func"](card)

        all_sources = SCRAPER_CONFIG.keys()
        prices = []
        urls = []
        display_data = {}
        for name in all_sources:
            if name in results:
                result = results[name]
                price, _, url = result
                prices.append((name, price))
                urls.append((name, url))
                if name == "Hareruya" and f"{name}_jpy" in results:
                    display_data[name] = f"{price:.2f} (¥{results[name+'_jpy']})"
                else:
                    display_data[name] = f"{price:.2f}"
            else:
                display_data[name] = "--"

        cheapest_price = min((p for _, p in prices if p > 0), default=0.0)
        cheapest_url = next((u for n, u in urls if n in results and abs(results[n][0] - cheapest_price) < 0.001), '')
        return (card, display_data, cheapest_price, cheapest_url, results)


    def open_cheapest_from_source(self, source):
        if not self.card_urls:
            messagebox.showinfo("No Results", "Please run a search first.")
            return
        opened = 0
        for card, data in self.card_urls.items():
            cheapest_url = data.get("Cheapest", "")
            source_url = data.get("URLs", {}).get(source, "")
            if source_url and source_url == cheapest_url:
                webbrowser.open_new_tab(source_url)
                opened += 1
        messagebox.showinfo("Done", f"Opened {opened} cheapest links from {source}.")

    def open_from_source(self, source):
        selected = self.tree.selection()
        if not selected:
            return
        row_id = selected[0]
        card_name = self.tree.item(row_id)['values'][0]
        url = self.card_urls.get(card_name, {}).get("URLs", {}).get(source, "")
        if url:
            webbrowser.open_new_tab(url)
        else:
            messagebox.showinfo("No Link", f"No {source} link available for {card_name}.")


    def open_all_cheapest_by_source(self):
        if not self.card_urls:
            messagebox.showinfo("No Results", "Please run a search first.")
            return
        opened = 0
        for source in SCRAPER_CONFIG:
            for card, data in self.card_urls.items():
                url = data.get("URLs", {}).get(source, "")
                price_str = data.get("Prices", {}).get(source, "--")
                try:
                    price = float(price_str)
                except:
                    price = 0.0
                if url and price > 0:
                    webbrowser.open_new_tab(url)
                    opened += 1
        messagebox.showinfo("Done", f"Opened {opened} total links from all sources.")

    def check_prices(self):
        self.tree.delete(*self.tree.get_children())
        self.card_urls.clear()
        self._results_cache.clear()
        self.total_label.config(text='Total: AU $0.00')
        input_text = self.text_input.get('1.0', tk.END)
        cards = parse_decklist_from_input(input_text)
        total = 0.0

        limiter = RateLimiter(3)  

        for i, card in enumerate(cards, start=1):
            if self.stop_flag:
                self.progress_label.config(text='Stopped.')
                break

            card_key = card.strip().lower()
            if card_key in self._results_cache:
                display_data, cheapest, url, results = self._results_cache[card_key]
            else:
                limiter.wait()
                card, display_data, cheapest, url, results = self.fetch_card_prices_parallel(card)
                self._results_cache[card_key] = (display_data, cheapest, url, results)

            row = [card]
            for source in SCRAPER_CONFIG:
                row.append(display_data.get(source, "--"))
            row.append(f"{cheapest:.2f}")
            # CK% comparison
            ck_usd = get_ck_price(card)
            if ck_usd and cheapest > 0:
                ratio = ck_usd / cheapest  # >1 means we're cheaper than CK
                pct = ratio * 100
                ck_display = f"{pct:.0f}%"
            else:
                ck_display = "--"
            row.append(ck_display)
            iid = self.tree.insert('', 'end', values=tuple(row))
            # Apply row tag for CK% colouring (done after insert)
            if ck_usd and cheapest > 0:
                self._apply_ck_tag(iid, ratio)

            self.card_urls[card] = {
                'Cheapest': url,
                'Prices': {source: display_data.get(source, "--") for source in SCRAPER_CONFIG},
                'URLs': {source: results.get(source, (0.0, "", ""))[2] if results.get(source) else "" for source in SCRAPER_CONFIG}
            }

            total += cheapest
            self.total_label.config(text=f'Total: AU ${total:.2f}')
            self.progress_label.config(text=f'Processing: {i}/{len(cards)}')
            self.root.update_idletasks()

        self.progress_label.config(text='Done' if not self.stop_flag else 'Stopped.')
        self.button.config(text='Search Prices', state='normal')
        self.hareruya_lang_dropdown.config(state='readonly')
        self.update_visible_columns()
        self.recalculate_cheapest_prices()

    def open_all_cheapest(self):
        if not self.card_urls:
            messagebox.showinfo("No Results", "Please run a search first.")
            return
        opened = 0
        for card, sources in self.card_urls.items():
            url = sources.get("Cheapest")
            if url:
                webbrowser.open_new_tab(url)
                opened += 1
        messagebox.showinfo("Done", f"Opened {opened} links in your browser.")

    def save_to_excel(self):
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        if not self.card_urls:
            messagebox.showinfo("No Data", "You must search prices before saving.")
            return

        wb = Workbook()

        # ── Sheet 1: Summary ─────────────────────────────────────────────────
        ws = wb.active
        ws.title = "Price Summary"

        FONT_NAME = "Arial"
        HDR_FILL  = PatternFill("solid", start_color="1F4E79")   # dark blue
        ALT_FILL  = PatternFill("solid", start_color="D9E1F2")   # light blue
        GRN_FILL  = PatternFill("solid", start_color="E2EFDA")
        RED_FILL  = PatternFill("solid", start_color="FCE4D6")
        NEU_FILL  = PatternFill("solid", start_color="FFFFFF")
        thin = Side(style="thin", color="AAAAAA")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        sources = list(SCRAPER_CONFIG.keys())
        headers = ["Card", "Cheapest (AU$)", "Cheapest Source"] +                   [f"{s} (AU$)" for s in sources] +                   ["CK Price (USD)", "CK Ratio", "CK% vs AUD"]

        # Write headers
        for col_idx, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=h)
            cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
            cell.fill = HDR_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        ws.row_dimensions[1].height = 30

        # Determine cheapest source per card
        def cheapest_source(card_name, cheapest_price_str):
            try:
                cheapest_val = float(str(cheapest_price_str).split()[0])
            except:
                return ""
            urls_data = self.card_urls.get(card_name, {})
            cheapest_url = urls_data.get("Cheapest", "")
            for src in sources:
                if urls_data.get("URLs", {}).get(src, "") == cheapest_url and cheapest_url:
                    return src
            return ""

        rows_data = []
        for row_id in self.tree.get_children():
            row = self.tree.item(row_id)["values"]
            card = row[0]
            cheapest_price_str = row[len(SCRAPER_CONFIG) + 1]
            try:
                cheapest_val = float(str(cheapest_price_str).split()[0])
            except:
                cheapest_val = 0.0
            src = cheapest_source(card, cheapest_price_str)
            cheapest_url = self.card_urls.get(card, {}).get("Cheapest", "")

            source_prices = {}
            for s in sources:
                p_str = self.card_urls.get(card, {}).get("Prices", {}).get(s, "--")
                try:
                    source_prices[s] = float(str(p_str).split()[0])
                except:
                    source_prices[s] = None

            ck_usd = get_ck_price(card)
            rows_data.append((card, cheapest_val, src, cheapest_url, source_prices, ck_usd))

        for r_idx, (card, cheapest_val, src, cheapest_url, source_prices, ck_usd) in enumerate(rows_data, 2):
            fill = ALT_FILL if r_idx % 2 == 0 else NEU_FILL

            col = 1
            # Card name with hyperlink to cheapest
            cell = ws.cell(row=r_idx, column=col, value=card)
            if cheapest_url:
                cell.hyperlink = cheapest_url
                cell.font = Font(name=FONT_NAME, size=10, color="0563C1", underline="single")
            else:
                cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = fill; cell.border = border
            col += 1

            # Cheapest price
            cell = ws.cell(row=r_idx, column=col, value=cheapest_val if cheapest_val > 0 else None)
            cell.number_format = "$#,##0.00"
            cell.font = Font(name=FONT_NAME, size=10, bold=True)
            cell.fill = fill; cell.border = border; cell.alignment = Alignment(horizontal="center")
            col += 1

            # Cheapest source
            cell = ws.cell(row=r_idx, column=col, value=src)
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = fill; cell.border = border; cell.alignment = Alignment(horizontal="center")
            col += 1

            # Per-source prices
            for s in sources:
                price = source_prices.get(s)
                src_url = self.card_urls.get(card, {}).get("URLs", {}).get(s, "")
                cell = ws.cell(row=r_idx, column=col, value=price)
                cell.number_format = "$#,##0.00"
                cell.font = Font(name=FONT_NAME, size=10,
                                 color="0563C1" if src_url else "000000",
                                 underline="single" if src_url else "none")
                if src_url:
                    cell.hyperlink = src_url
                if price and price == cheapest_val and cheapest_val > 0:
                    cell.fill = GRN_FILL
                else:
                    cell.fill = fill
                cell.border = border; cell.alignment = Alignment(horizontal="center")
                col += 1

            # CK USD price
            cell = ws.cell(row=r_idx, column=col, value=ck_usd)
            cell.number_format = "$#,##0.00"
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = fill; cell.border = border; cell.alignment = Alignment(horizontal="center")
            col += 1

            # CK ratio and % (calculated in Python to avoid formula escaping issues)
            ratio_fill = fill
            ratio_val = None
            pct_val = None
            if ck_usd and cheapest_val > 0:
                ratio_val = round(ck_usd / cheapest_val, 4)
                pct_val = round(ratio_val - 1, 4)
                if ratio_val >= 1.2:
                    ratio_fill = GRN_FILL
                elif ratio_val <= 0.8:
                    ratio_fill = RED_FILL

            cell = ws.cell(row=r_idx, column=col, value=ratio_val)
            cell.number_format = '0.00"x"'
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = ratio_fill; cell.border = border; cell.alignment = Alignment(horizontal="center")
            col += 1

            cell = ws.cell(row=r_idx, column=col, value=pct_val)
            cell.number_format = "+0%;-0%;0%"
            cell.font = Font(name=FONT_NAME, size=10)
            cell.fill = ratio_fill; cell.border = border; cell.alignment = Alignment(horizontal="center")

        # Total row
        total_row = len(rows_data) + 2
        ws.cell(row=total_row, column=1, value="TOTAL").font = Font(name=FONT_NAME, bold=True, size=10)
        total_cell = ws.cell(row=total_row, column=2,
                             value=f"=SUM(B2:B{total_row-1})")
        total_cell.number_format = "$#,##0.00"
        total_cell.font = Font(name=FONT_NAME, bold=True, size=11)
        total_cell.fill = PatternFill("solid", start_color="1F4E79")
        total_cell.font = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=11)
        total_cell.border = border

        # Column widths
        ws.column_dimensions["A"].width = 26
        ws.column_dimensions["B"].width = 14
        ws.column_dimensions["C"].width = 14
        for i in range(4, 4 + len(sources)):
            ws.column_dimensions[get_column_letter(i)].width = 13
        ck_start = 4 + len(sources)
        ws.column_dimensions[get_column_letter(ck_start)].width = 14
        ws.column_dimensions[get_column_letter(ck_start+1)].width = 10
        ws.column_dimensions[get_column_letter(ck_start+2)].width = 10
        ws.freeze_panes = "A2"

        # ── Sheet 2: Missing Cards ───────────────────────────────────────────
        ws2 = wb.create_sheet("Missing Cards")
        ws2.append(["Card Name"])
        ws2["A1"].font = Font(name=FONT_NAME, bold=True, color="FFFFFF")
        ws2["A1"].fill = HDR_FILL
        ws2["A1"].border = border
        missing = [self.missing_listbox.get(i) for i in range(self.missing_listbox.size())]
        for card in missing:
            ws2.append([card])
        ws2.column_dimensions["A"].width = 30

        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"MTG_Price_Report_{timestamp}.xlsx"
        filepath = os.path.join(os.path.expanduser("~/Downloads"), filename)
        try:
            wb.save(filepath)
            messagebox.showinfo("Saved", f"Excel saved to:\n{filepath}")
        except Exception as e:
            messagebox.showerror("Error", f"Failed to save:\n{e}")

    def on_click(self, event):
        selected = self.tree.selection()
        if not selected:
            return

        row_id = selected[0]

        if self.last_selected_row == row_id:
            item = self.tree.item(row_id)
            card_name = item['values'][0]
            urls = self.card_urls.get(card_name, {})
            url = urls.get("Cheapest")
            if url:
                webbrowser.open_new_tab(url)
        else:
            self.last_selected_row = row_id


    def refresh_deck_dropdown(self):
        names = list(self.deck_cache.keys())
        self.deck_dropdown["values"] = names

        if not names:
            self.deck_dropdown.set("Select saved deck")
            self.deck_dropdown.configure(foreground="grey")
        else:
            current = self.deck_var.get()
            if current not in names:
                self.deck_dropdown.set("Select saved deck")
                self.deck_dropdown.configure(foreground="grey")



    def save_deck(self):
        url = self.url_entry.get().strip()
        deck_text = self.text_input.get("1.0", tk.END).strip()
        if not url or not deck_text:
            messagebox.showwarning("Error", "Need both a deck URL and decklist to save.")
            return

        deck_name = None

        if "archidekt.com/decks/" in url:
            try:
                slug = url.rstrip("/").split("/")[-1]
                deck_name = slug.replace("_", " ").title()
            except Exception as e:
                print(f"[Archidekt name parse error] {e}")

        if not deck_name and hasattr(self, "parsed_deck_name") and self.parsed_deck_name:
            deck_name = self.parsed_deck_name

        if not deck_name and "moxfield.com" in url:
            try:
                resp = requests.get(url, timeout=15)
                soup = BeautifulSoup(resp.text, "html.parser")
                tag = soup.select_one("span.deckHeader_deckName__OlKwW")
                if tag:
                    deck_name = tag.get_text(strip=True)
            except Exception as e:
                print(f"[Deck name fetch error] {e}")

        if not deck_name:
            deck_name = f"Deck {len(self.deck_cache)+1}"

        self.deck_cache[deck_name] = {
            "url": url,
            "decklist": deck_text
        }
        save_deck_cache(self.deck_cache)
        self.refresh_deck_dropdown()
        messagebox.showinfo("Saved", f"Deck saved as '{deck_name}'")

    def delete_deck(self):
        name = self.deck_var.get()
        if name in self.deck_cache:
            del self.deck_cache[name]
            save_deck_cache(self.deck_cache)
            self.refresh_deck_dropdown()

            self.url_entry.delete(0, tk.END)
            self.url_entry.insert(0, "Paste a deck link")
            self.url_entry.config(fg="grey")

            self.text_input.delete("1.0", tk.END)

            self.deck_dropdown.set("Select saved deck")
            self.deck_dropdown.configure(foreground="grey")

            messagebox.showinfo("Deleted", f"Removed deck '{name}'")
        else:
            messagebox.showwarning("Error", "No saved deck selected to delete.")


    def load_saved_deck(self, event=None):
        name = self.deck_var.get()
        if name and name in self.deck_cache:
            data = self.deck_cache[name]

            self.url_entry.delete(0, tk.END)
            self.url_entry.insert(0, data.get("url", ""))
            self.url_entry.config(fg="black")

            self.text_input.delete("1.0", tk.END)
            self.text_input.insert(tk.END, data.get("decklist", ""))

            self.deck_dropdown.configure(foreground="black")


if __name__ == '__main__':
    root = TkinterDnD.Tk()  
    app = MTGScraperGUI(root)
    root.mainloop()


